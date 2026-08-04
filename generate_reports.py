#!/usr/bin/env python3
"""
Hiktaos Report Card Generator - preserves all template formatting
"""

import csv, re, os, shutil, zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

GRADE_TABLE = [
    (75,100,"A1","Excellent"),(70,74,"B2","Very Good"),
    (65,69,"B3","Good"),(60,64,"C4","Credit"),
    (55,59,"C5","Credit"),(50,54,"C6","Pass"),
    (45,49,"D7","Fair"),(40,44,"E8","Weak"),
    (0,39,"F9","Fail"),
]

THIN = Border(left=Side(style='thin'),right=Side(style='thin'),
              top=Side(style='thin'),bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center',vertical='center',wrap_text=True)
LEFT = Alignment(horizontal='left',vertical='center',wrap_text=True)
F9 = Font(name='Arial',size=9)
F9B = Font(name='Arial',size=9,bold=True)
F11B = Font(name='Arial',size=11,bold=True)
F12B = Font(name='Arial',size=12,bold=True)
F14G = Font(name='Arial',size=14,bold=True,color='006400')

PROT = ['xl/media/image1.jpeg','xl/drawings/drawing1.xml',
        'xl/drawings/_rels/drawing1.xml.rels',
        'xl/worksheets/_rels/sheet1.xml.rels']

def grd(t):
    t=max(0,min(100,t))
    for l,h,g,r in GRADE_TABLE:
        if l<=t<=h: return g,r
    return "F9","Fail"

def save(fp, tmpl, mod):
    shutil.copy2(tmpl, fp)
    wb = load_workbook(fp)
    mod(wb)
    t = fp+'.tmp'
    wb.save(t); wb.close()
    with zipfile.ZipFile(t) as z: o = {n:z.read(n) for n in z.namelist()}
    with zipfile.ZipFile(tmpl) as z:
        for n in PROT:
            if n in z.namelist(): o[n] = z.read(n)
    with zipfile.ZipFile(fp,'w',zipfile.ZIP_DEFLATED) as z:
        for n in sorted(o): z.writestr(n,o[n])
    os.remove(t)

def run(cfg):
    cn, nx, tot, tmpl, data, csv_m, suf, sm, ren, summ_start, max_subj_row = cfg
    base = os.path.dirname(__file__)
    out = os.path.join(base, 'report_cards')
    os.makedirs(os.path.join(out, cn), exist_ok=True)
    ORD = ["","1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th",
           "11th","12th","13th","14th","15th","16th","17th","18th"]
    
    stu = []
    if csv_m:
        with open(os.path.join(base, data)) as f:
            for r in csv.DictReader(f):
                s = {'n': r.get('Names', r.get('Student', r.get('Student Name', '')))}
                for p, _ in sm:
                    m = re.match(r'(\d+)', str(r.get(p,"0")))
                    v = int(m.group(1)) if m else 0
                    s[p] = v
                # Add free YORUBA scores for SS1 or SS2
                if cn == "SS 1" and s['n'] in SS1_YORUBA:
                    s['YORUBA'] = SS1_YORUBA[s['n']]
                if cn == "SS 2" and s['n'] in SS2_YORUBA:
                    s['YORUBA'] = SS2_YORUBA[s['n']]
                # SS2 Computer: read CA/Exam from CSV
                if cn == "SS 2":
                    s['Computer_Total'] = int(r.get('Computer_Total', '0'))
                    s['Computer_CA'] = int(r.get('Computer_CA', '0'))
                    s['Computer_Exam'] = int(r.get('Computer_Exam', '0'))
                stu.append(s)
    else:
        wb = load_workbook(os.path.join(base, data), data_only=True)
        sheet_name = 'Scores' if 'Scores' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        hd = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=1, column=c).value
            if v: hd[v] = c
        for r in range(2, ws.max_row+1):
            nm = ws.cell(row=r, column=hd.get('Student Name', 2)).value
            if not nm: continue
            s = {'n': nm}
            for p, rn in sm:
                cc = hd.get(f'{p}{suf}')
                ec = hd.get(f'{p}_Exam')
                if cc and ec:
                    cv = ws.cell(row=r, column=cc).value or 0
                    ev = ws.cell(row=r, column=ec).value or 0
                    s[p] = cv + ev
                elif cc and not ec:
                    cv = ws.cell(row=r, column=cc).value or 0
                    s[p] = cv
            # Add free YORUBA scores for SS1 or SS2
            if cn == "SS 1" and nm in SS1_YORUBA:
                s['YORUBA'] = SS1_YORUBA[nm]
            if cn == "SS 2" and nm in SS2_YORUBA:
                s['YORUBA'] = SS2_YORUBA[nm]
            # SS2 Computer: read CA/Exam from data
            if cn == "SS 2":
                s['Computer_Total'] = s.get('Computer_Total', 0)
                s['Computer_CA'] = s.get('Computer_CA', 0)
                s['Computer_Exam'] = s.get('Computer_Exam', 0)
            stu.append(s)
    
    if not stu: return
    
    tl = [(sum(s.get(p,0) for p,_ in sm), s['n']) for s in stu]
    tl.sort(key=lambda x: -x[0])
    rk = {}; ri = 0; pr = None
    for i,(t,n) in enumerate(tl):
        if pr is not None and t < pr: ri = i
        rk[n] = ri + 1; pr = t
    
    sr = {}
    for p, _ in sm:
        sc = [(s.get(p,0), s['n']) for s in stu]
        sc.sort(key=lambda x: -x[0])
        sr2 = {}; ri = 0; pr = None
        for i,(t,n) in enumerate(sc):
            if pr is not None and t < pr: ri = i
            sr2[n] = ri + 1; pr = t
        sr[p] = sr2
    
    used = {r for _,r in sm}
    empty = set(range(28, max_subj_row+1)) - used
    
    print(f"\n📚 {cn} ({len(stu)} students, {len(sm)} subjects)")
    
    for s in stu:
        nm = s['n']; rnk = rk[nm]
        sf = nm.replace(' ','_').replace("'",'').replace('/','_')
        fp = os.path.join(out, cn, f"Report_Card_{sf}.xlsx")
        
        def mod(wb):
            ws = wb['Report Sheet']
            ws['D16'].value = nm; ws['D16'].font = F12B
            ws['D20'].value = cn; ws['D20'].font = F12B; ws['D20'].alignment = LEFT
            ws['M19'].value = tot; ws['M19'].font = F12B; ws['M19'].alignment = CENTER
            ws['M21'].value = tot; ws['M21'].font = F12B; ws['M21'].alignment = CENTER
            
            for rn, nn in ren.items():
                ws.cell(row=rn, column=2).value = nn
                ws.cell(row=rn, column=2).font = F9; ws.cell(row=rn, column=2).alignment = LEFT; ws.cell(row=rn, column=2).border = THIN
            
            for e in empty:
                for c in [7,8,10,12,14,16]:
                    try: ws.cell(row=e, column=c).value = ""
                    except: pass
            
            for r in range(60, 75):
                try:
                    if ws.cell(row=r, column=19).value and 'NEXT' in str(ws.cell(row=r, column=19).value):
                        ws.cell(row=r, column=22).value = nx
                        ws.cell(row=r, column=22).font = F12B; ws.cell(row=r, column=22).alignment = CENTER
                        break
                except: pass
            
            tt = 0; cr = 0; ct = 0
            for p, rn in sm:
                tv = s.get(p, 0)
                if tv == 0:
                    for c in [7,8,10,12,14,16]:
                        try: ws.cell(row=rn, column=c).value = ""
                        except: pass
                    continue
                ct += 1; tt += tv
                gl, _ = grd(tv)
                if gl in ("A1","B2","B3","C4","C5","C6"): cr += 1
                ca = tv * 40 / 100
                gv = round(ca / 2)
                hv = round(ca - gv)
                jv = tv - gv - hv
                if jv > 60: jv = 60; hv = tv - gv - jv
                # Computer: split CA into G+H, Exam goes to J
                if p == 'Computer_Total':
                    comp_ca = int(s.get('Computer_CA', 0))
                    comp_ex = int(s.get('Computer_Exam', 0))
                    gv = round(comp_ca / 2)
                    hv = comp_ca - gv
                    jv = comp_ex
                    ws.cell(row=rn, column=7).value = gv
                    ws.cell(row=rn, column=8).value = hv
                    ws.cell(row=rn, column=10).value = jv
                else:
                    ws.cell(row=rn, column=7).value = gv
                    ws.cell(row=rn, column=8).value = hv
                    ws.cell(row=rn, column=10).value = jv
                rr = sr[p].get(nm, 0)
                rs = ORD[rr] if 1 <= rr <= len(stu) else ""
                ws.cell(row=rn, column=16).value = rs
                ws.cell(row=rn, column=16).font = F9; ws.cell(row=rn, column=16).alignment = CENTER; ws.cell(row=rn, column=16).border = THIN
                for c in [7,8,10]:
                    cl = ws.cell(row=rn, column=c)
                    cl.font = F9; cl.alignment = CENTER; cl.border = THIN
            
            avg = round(tt/max(ct,1),1)
            pct = round(tt/max(ct*100,1)*100,2)
            ov, _ = grd(round(pct))
            gp = {"A1":4.0,"B2":3.5,"B3":3.0,"C4":2.5,"C5":2.0,"C6":1.5,"D7":1.0,"E8":0.5,"F9":0.0}.get(ov,0.0)
            ps = ORD[rnk] if 1 <= rnk <= 5 else ""
            
            # LEFT summary at configurable start row
            for i,(r,v) in enumerate([(summ_start,tt),(summ_start+1,avg),(summ_start+2,f"{pct}%"),
                         (summ_start+3,ov),(summ_start+4,ps),(summ_start+5,tot),
                         (summ_start+6,cr)]):
                try: ws.cell(row=r,column=7).value = v; ws.cell(row=r,column=7).font = F11B
                except: pass
            # GPA at summ_start+7
            try: ws.cell(row=summ_start+7, column=7).value = gp; ws.cell(row=summ_start+7, column=7).font = F11B
            except: pass
            
            ws.cell(row=20, column=13).value = ps
            ws.cell(row=20, column=13).font = F12B; ws.cell(row=20, column=13).alignment = CENTER
            
            for r,v in [(38,tt),(39,avg),(40,f"{pct}%"),(41,ov),(42,ps),(43,tot),(44,cr),(45,gp)]:
                try: ws.cell(row=r,column=24).value = v; ws.cell(row=r,column=24).font = F11B
                except: pass
            
            for r in range(60, 70):
                try:
                    if ws.cell(row=r,column=20).value == "PROMOTED":
                        ws.cell(row=r,column=20).value = "✔  PROMOTED"
                        ws.cell(row=r,column=20).font = F14G; ws.cell(row=r,column=20).alignment = CENTER
                except: pass
        
        save(fp, os.path.join(base, tmpl), mod)
        print(f"    ✅ {nm}")

# YORUBA free scores for SS1 (max ≤ 85, highest = Owodunni, Mateen, Paul)
SS1_YORUBA = {
    'Owodunni Hikimah': 85, 'Arolu Mateen': 84, 'James Paul': 83,
    'Nsini Destiny': 82, 'Ademu Samuel': 80, 'Kolawole Samuel': 78,
    'Obeya Emmanuella': 75, 'James Peter': 72, 'Gbadebo Esther': 70,
    'Alex Favour': 68, 'Sulaimon Monsurah': 65, 'Adebayo Khadijat': 62,
    'Yusuf Rofia': 58, 'Faronbi Oyinkansola': 55, 'Lowo Adunni': 50,
    'James Hannah': 45,
}

# YORUBA free scores for SS2 (Barakat and Balogun highest)
SS2_YORUBA = {
    'Shakiru Baraka': 85, 'Balogun Khadijat': 84,
    'Adewale Abidemi': 83, 'Micheal Daniel': 82,
    'Awolola Boluwatife': 80, 'Akorede Fathia': 75,
    'Daniel Deborah': 72, 'Egbordor Joseph': 68,
    'Alex Esther': 65, 'Jonah Happiness': 62,
    'Akinbile Fausat': 55,
}

# CFGS: (name, next, total, template, data, csv, suffix, [(prefix,row),...], {renames}, summary_start, max_subj_row)
CFGS = [
    ("Primary 4", "Primary 5", 10, "Hiktaos_Primary_Final.xlsx", "academic-performance(2).csv", True, "",
     [("Eng",40),("Maths",39),("Basic-S",35),("Social Studies",28),("Civic",29),
      ("CCA",38),("PHE",36),("Agric",33),("Home-E",34),("CRS",30),
      ("Security",31),("History",32),("Quantitative",42),("Verbal",41),("ICT",37),("Vocational",43)],
     {}, 44, 43),
    ("Primary 3", "Primary 4", 12, "Hiktaos_Primary_Final.xlsx", "Primary_3_Hiktaos_Student_Scores.xlsx", False, "_Test",
     [("Social Studies",28),("Civic",29),("IRK_CRK",30),("Security",31),("History",32),
      ("Basic Science",35),("PHE",36),("ICT",37),("CCA",38),
      ("Mathematics",39),("English",40),("Verbal",41),("Quantitative",42),("Vocational",43)],
     {}, 44, 43),
    ("Basic 2", "Basic 3", 6, "Hiktaos_Primary_Final.xlsx", "Primary_2_Hiktaos_Student_Scores.xlsx", False, "_CA",
     [("Social Studies",28),("Civic Education",29),("CRS",30),("Secondary Education",31),
      ("History",32),("Basic Science",35),("Physical Education",36),("Computer Studies",37),
      ("Cultural Art",38),("Mathematics",39),("English",40),("Verbal Reasoning",41),
      ("Quantitative Reasoning",42),("Vocational",43)],
     {31:"Secondary Education",36:"Physical Education",37:"Computer Studies",38:"Cultural Art"}, 44, 43),
    ("Primary 1", "Primary 2", 15, "Hiktaos_Primary_Final.xlsx", "Primary_1_Hiktaos_Student_Scores.xlsx", False, "_CA",
     [("Social Studies",28),("Civic Education",29),("CRS",30),("Security Education",31),
      ("History",32),("Basic Science",35),("Physical Education",36),("Computer Studies",37),
      ("Cultural Art",38),("Mathematics",39),("English",40),("Verbal Reasoning",41),
      ("Quantitative Reasoning",42),("Vocational",43)],
     {36:"Physical Education",37:"Computer Studies",38:"Cultural Art"}, 44, 43),
    ("JSS 2", "JSS 3", 13, "Hiktaos_JSS_Final.xlsx", "JSS_2_Hiktaos_Student_Scores.xlsx", False, "_CA",
     [("English",28),("Mathematics",29),("Security Education",30),("Basic Science",31),
      ("Social Studies",32),("Civic Education",33),("Creative Art",34),("Agricultural Science",35),
      ("Physical Education",36),("Business Studies",37),("Home Economics",38),
      ("Basic Technology",39),("ICT",40),("CRK_IRS",41)],
     {30:"Security Education",34:"Creative Art",36:"Physical Education",40:"ICT",41:"CRK_IRS"}, 44, 43),
     
    # SS1 - single scores from named columns, summary at row 48, Yoruba free scores
    ("SS 1", "SS 2", 16, "Hiktaos_SS_Final.xlsx", "SS1_Third_Term_Score_Sheet.xlsx", False, "",
     [("MTH",29),("ENG",28),("AGR",36),("ECO",39),("CRK",46),
      ("CHM",32),("PHY",31),("CIV",30),("GOV",43),("COM",35),
      ("BIO",33),("LIT",42),("COMMERCE",38),("ACCOUNTING",37),
      ("YORUBA",47)],
     {}, 48, 47),
     
    # SS2 - same as SS1, uses _CA/_Exam columns, Yoruba free scores
    ("SS 2", "SS 3", 11, "Hiktaos_SS_Final.xlsx", "student_scores.csv", True, "",
     [("English_Total",28),("Maths_Total",29),("Economics_Total",39),("Agric_Total",36),
      ("Commerce_Total",38),("Account_Total",37),("Chemistry_Total",32),("Physics_Total",31),
      ("Literature_Total",42),("CRS_IRS_Total",46),("Biology_Total",33),("Civic_Total",30),
      ("Government_Total",43),("Computer_Total",35),("YORUBA",47)],
     {}, 48, 47),
]

def main():
    for c in CFGS:
        run(c)

if __name__ == '__main__':
    main()
