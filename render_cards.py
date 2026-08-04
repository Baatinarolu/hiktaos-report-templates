#!/usr/bin/env python3
"""
Render Hiktaos report-card .xlsx files to JPEG and PDF using a from-scratch
renderer (openpyxl reads real styles -> Pillow draws).

Used because the sandbox has no LibreOffice / fonts / ghostscript and only
PyPI is reachable, so a faithful external converter cannot be installed.

Reads actual fills / borders / fonts / merges / images from the workbook and
re-draws them. Formula cells (Total =SUM, Grade =IF) are computed here since
openpyxl does not cache their values.
"""
import os, re, sys, io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "report_cards")
OUT_PDF = os.path.join(BASE, "report_cards_pdf")
OUT_JPG = os.path.join(BASE, "report_cards_jpeg")

DEJAVU = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
DEJAVU_B = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
SCALE = 2                      # render at 2x for crisp output (~192 dpi)
DEF_COL = 8.43                 # default Excel column width (chars)
DEF_ROW = 15.0                 # default row height (points)

# theme index -> RGB (matches this workbook's theme1.xml, with lt1/dk1 swap)
THEME = {0:(255,255,255),1:(0,0,0),2:(238,236,225),3:(31,73,125),
         4:(79,129,189),5:(192,80,77),6:(155,187,89),7:(128,100,162),
         8:(75,172,198),9:(247,150,70)}
# standard indexed palette (subset)
IDX = {0:(0,0,0),1:(255,255,255),8:(0,0,0),9:(255,255,255),64:(0,0,0),65:(255,255,255)}

# emoji / unsupported glyph substitution (DejaVu has no colour emoji)
EMOJI_MAP = {
    "\U0001F4CD":"",   # pushpin
    "\U0001F4DE":"Tel ",# telephone
    "\U0001F4E7":"", "\u2709":"", "\u2709\ufe0f":"",  # envelope/email
    "\U0001F310":"Web ",# globe
    "\U0001F499":"", "\U0001F49B":"", "\U0001F49C":"",  # hearts
    "\u2699":"", "\u2699\ufe0f":"",                      # gear
    "\U0001F393":"",                                     # graduation cap
    "\u2714":"\u2713",                                  # heavy check -> check
    "\u2714\ufe0f":"\u2713",
    "\u2B50":"\u2605",                                  # star
    "\U0001F3C6":"",                                    # trophy
    "\U0001F4D6":"", "\U0001F4DA":"",                   # books
    "\U0001F50D":"",                                    # search
    "\u2728":"",                                        # sparkles
}

_font_cache = {}
def font_for(name, size_pt, bold, italic):
    key = (bool(bold), bool(italic), round(size_pt*4/3*SCALE))
    if key in _font_cache: return _font_cache[key]
    path = DEJAVU_B if (bold or (name and "Black" in name)) else DEJAVU
    try:
        f = ImageFont.truetype(path, round(size_pt*4/3*SCALE))
    except Exception:
        f = ImageFont.truetype(DEJAVU, round(size_pt*4/3*SCALE))
    _font_cache[key] = f
    return f

def fix_text(s):
    if s is None: return ""
    s = str(s)
    for k,v in EMOJI_MAP.items():
        s = s.replace(k,v)
    return s

def col_width(ws, col):   # 1-based col -> px @96dpi
    L = get_column_letter(col)
    d = ws.column_dimensions.get(L)
    w = d.width if d and d.width else DEF_COL
    return round(w*7+5)

def row_height(ws, row):  # row -> px @96dpi
    d = ws.row_dimensions.get(row)
    h = d.height if d and d.height else DEF_ROW
    return round(h*4/3)

def color_of(c):
    if c is None: return None
    t = getattr(c,"type",None)
    if t == "rgb" and isinstance(c.rgb,str):
        return hexrgb(c.rgb)
    if t == "theme":
        return THEME.get(int(c.theme),(0,0,0))
    if t == "indexed":
        return IDX.get(int(c.indexed),(0,0,0))
    if isinstance(c.rgb,str):
        return hexrgb(c.rgb)
    return None

def hexrgb(h):
    h = h[-6:]
    try: return tuple(int(h[i:i+2],16) for i in (0,2,4))
    except Exception: return None

GRADE = [(75,"A1"),(70,"B2"),(65,"B3"),(60,"C4"),(55,"C5"),
         (50,"C6"),(45,"D7"),(40,"E8"),(0,"F9")]
def grade_letter(v):
    try: v=float(v)
    except: return ""
    for lo,g in GRADE:
        if v>=lo: return g
    return "F9"

def ref_to_num(ref):  # "G28" -> (col,row)
    m=re.match(r"([A-Z]+)(\d+)", ref)
    col=0
    for ch in m.group(1): col=col*26+(ord(ch)-64)
    return col, int(m.group(2))

def eval_formula(val, getval):
    s=val.strip().lstrip("=")
    if s.startswith("SUM(") or "SUM(" in s:
        refs=re.findall(r"[A-Z]+\d+", s)
        tot=0
        for r in refs:
            v=getval(r)
            try: tot+=float(v)
            except: pass
        return int(tot) if float(tot).is_integer() else tot
    if s.startswith("IF(") and ('"A1"' in s or '"F9"' in s):
        # grade formula referencing an L cell
        refs=re.findall(r"[A-Z]+\d+", s)
        for r in refs:
            v=getval(r)
            try:
                return grade_letter(float(v))
            except: pass
        return ""
    return val

# ---------- main render ----------
def render(path):
    wb = load_workbook(path)
    ws = wb.active
    ncol = ws.max_column
    nrow = ws.max_row
    for mr in ws.merged_cells.ranges:
        ncol=max(ncol,mr.max_col); nrow=max(nrow,mr.max_row)

    # cumulative base coordinates (px @96dpi), sized generously
    xs=[0]*(ncol+4)
    for c in range(1,ncol+3): xs[c]=xs[c-1]+col_width(ws,c)
    ys=[0]*(nrow+4)
    for r in range(1,nrow+3): ys[r]=ys[r-1]+row_height(ws,r)
    W=xs[ncol]; H=ys[nrow]
    # pad so any coordinate access is safe
    xs=xs+[xs[-1]]*50
    ys=ys+[ys[-1]]*50

    # merged ranges: map every covered coord -> owner; list of ranges
    owned={}        # (r,c)-> (r0,c0,r1,c1)
    ranges=[]
    for mr in ws.merged_cells.ranges:
        r0,c0,r1,c1 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        ranges.append((r0,c0,r1,c1))
        for r in range(r0,r1+1):
            for c in range(c0,c1+1):
                owned[(r,c)]=(r0,c0,r1,c1)

    def getval(coord):
        try:
            cell=ws[coord]
            v=cell.value
            if isinstance(v,str) and v.startswith("="):
                return eval_formula(v, getval)
            return v
        except Exception:
            return 0

    img=Image.new("RGB",(W*SCALE,H*SCALE),"white")
    d=ImageDraw.Draw(img)

    def px(x): return x*SCALE

    # --- fills (merged-aware) ---
    drawn_fill=set()
    for r in range(1,nrow+1):
        for c in range(1,ncol+1):
            cell=ws.cell(row=r,column=c)
            if (r,c) in owned:
                owner=owned[(r,c)]
                if owner in drawn_fill: continue
                drawn_fill.add(owner)
                r0,c0,r1,c1=owner
                cell=ws.cell(row=r0,column=c0)
            fill=cell.fill
            col=None
            if fill and fill.patternType=="solid":
                col=color_of(fill.fgColor) or color_of(fill.bgColor)
            if col:
                d.rectangle([px(xs[c0-1]),px(ys[r0-1]),px(xs[c1])-1,px(ys[r1])-1], fill=col)

    # --- borders ---
    def border_width(sty):
        return {"thin":max(1,int(0.8*SCALE)),"medium":max(1,int(1.6*SCALE)),
                "thick":max(2,int(2.2*SCALE))}.get(sty,0)
    for r in range(1,nrow+1):
        for c in range(1,ncol+1):
            if (r,c) in owned and owned[(r,c)]!=(r,c):  # interior of merge, skip
                continue
            cell=ws.cell(row=r,column=c)
            b=cell.border
            # merged: use the merge's bounding box edges; style from owner borders
            if (r,c) in owned:
                r0,c0,r1,c1=owned[(r,c)]
            else:
                r0=c0=r; r1=c1=c; cc=c
            x0=px(xs[c0-1]); x1=px(xs[c1]); y0=px(ys[r0-1]); y1=px(ys[r1])
            def edge(side, x1a,y1a,x2a,y2a):
                sty=getattr(b,side)
                if sty and sty.style:
                    w=border_width(sty.style); col=color_of(sty.color) or (0,0,0)
                    d.line([x1a,y1a,x2a,y2a], fill=col, width=w)
            edge("top",    x0,y0,x1,y0)
            edge("bottom", x0,y1,x1,y1)
            edge("left",   x0,y0,x0,y1)
            edge("right",  x1,y0,x1,y1)

    # --- text ---
    def draw_text(cell, r, c, r0,c0,r1,c1):
        v=cell.value
        if v is None: return
        if isinstance(v,str) and v.startswith("="):
            v=eval_formula(v, getval)
        if v=="" or v is None: return
        txt=fix_text(v)
        if isinstance(v,float):
            txt=str(int(v)) if v.is_integer() else str(v)
        elif isinstance(v,int):
            txt=str(v)
        f=cell.font
        color=color_of(f.color) or (0,0,0)
        font=font_for(f.name, f.size or 11, f.bold, f.italic)
        al=cell.alignment
        ha=al.horizontal or ("left" if c0!=1 else "center")
        # available box
        bx0,bx1=px(xs[c0-1])+SCALE, px(xs[c1])-SCALE
        by0,by1=px(ys[r0-1])+1, px(ys[r1])-1
        boxw=bx1-bx0; boxh=by1-by0
        # wrap into lines
        words=txt.split(" ") if " " in txt else [txt]
        lines=[]; cur=""
        def wlen(s): return font.getlength(s)
        for w in words:
            trial=(cur+" "+w).strip()
            if wlen(trial)<=boxw or not cur:
                cur=trial
            else:
                lines.append(cur); cur=w
        if cur: lines.append(cur)
        if not lines: lines=[txt]
        line_h=font.size
        total_h=line_h*len(lines)
        # vertical
        if al.vertical=="top": ty=by0
        elif al.vertical=="bottom": ty=by1-total_h
        else: ty=by0+(boxh-total_h)//2
        for ln in lines:
            lw=wlen(ln)
            if ha=="center": tx=bx0+(boxw-lw)//2
            elif ha=="right": tx=bx1-lw
            else: tx=bx0
            d.text((tx,ty), ln, fill=color, font=font)
            ty+=line_h

    for r in range(1,nrow+1):
        for c in range(1,ncol+1):
            if (r,c) in owned and owned[(r,c)]!=(r,c): continue
            cell=ws.cell(row=r,column=c)
            if (r,c) in owned:
                r0,c0,r1,c1=owned[(r,c)]
            else:
                r0=c0=r; r1=c1=c
            draw_text(cell,r,c,r0,c0,r1,c1)

    # --- images (logo) ---
    for im in ws._images:
        a=im.anchor
        try:
            fr=a._from; to=a.to
            def cp(col,off):
                return (xs[col]+off/9525)
            x0=cp(fr.col,fr.colOff); y0=cp(fr.row,fr.rowOff)
            x1=cp(to.col,to.colOff); y1=cp(to.row,to.rowOff)
            data=im._data() if callable(getattr(im,"_data",None)) else im.ref and open(im.ref,'rb').read()
            if data:
                logo=Image.open(io.BytesIO(data)).convert("RGBA")
                logo=logo.resize((max(1,int((x1-x0)*SCALE)),max(1,int((y1-y0)*SCALE))))
                img.paste(logo,(int(x0*SCALE),int(y0*SCALE)),logo)
        except Exception as e:
            print("  img warn:",e)

    return img

def main():
    os.makedirs(OUT_PDF, exist_ok=True)
    os.makedirs(OUT_JPG, exist_ok=True)
    # collect cards from per-class subfolders: (path, class_folder)
    files = []
    for cls in sorted(os.listdir(SRC)):
        cd = os.path.join(SRC, cls)
        if not os.path.isdir(cd):
            continue
        for f in sorted(os.listdir(cd)):
            if f.startswith("Report_Card_") and f.endswith(".xlsx"):
                files.append((os.path.join(cd, f), cls))
    if len(sys.argv) > 1:   # optional filter substring for testing
        files = [(p, c) for p, c in files if sys.argv[1] in p] or files[:1]
    print("Rendering %d cards..." % len(files))
    for i, (fp, cls) in enumerate(files, 1):
        img = render(fp)
        stem = os.path.basename(fp)[:-5]
        jpg_dir = os.path.join(OUT_JPG, cls); os.makedirs(jpg_dir, exist_ok=True)
        pdf_dir = os.path.join(OUT_PDF, cls); os.makedirs(pdf_dir, exist_ok=True)
        img.save(os.path.join(jpg_dir, stem + ".jpg"), "JPEG", quality=82)
        small = img.resize((int(img.size[0]*0.7), int(img.size[1]*0.7)),
                           Image.LANCZOS)
        small.save(os.path.join(pdf_dir, stem + ".pdf"), "PDF", resolution=150.0)
        if i % 10 == 0 or i == len(files):
            print("  %d/%d [%s] %s" % (i, len(files), cls, stem))
    print("Done. ->", OUT_JPG, "and", OUT_PDF)

if __name__=="__main__":
    main()
