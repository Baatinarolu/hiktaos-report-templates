#!/usr/bin/env python3
"""
Hiktaos JSS 1 Report Card Generator.

Data source : JSS1 results.xlsx  (sheet 'Scores'; columns
              '<Subject>_Test', '<Subject>_Exam', '<Subject>_Total')
Template    : Hiktaos_JSS_Final.xlsx  (logo at xl/media/image1.jpeg)

Template score columns (subject rows 28-42):
    G  = CA 1
    H  = Assignment        (both under CONTINUOUS ASSESSMENT)
    J  = Examination
    L  = Total  -> formula =SUM(G,H,J)   (== Test + Exam)
    N  = Grade (formula already in template)
    P  = Position in subject

Per the standing instruction, the CA (the '_Test' value) is split across the
two CA boxes and the exam goes in the exam box:
    G = round(Test / 2)
    H = Test - G
    J = Exam
The card's displayed Total is therefore Test + Exam (matches the data's
'_Total' column exactly).

Subject mapping: 11 JSS1 subjects fit the template rows directly; Security
Education and History (not in the template) take the unused Yoruba(30) and
Business Studies(37) slots and are relabelled. CRS/IRS(41) and French(42)
are cleared (not offered).
"""
import os, re, shutil, zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE, "Hiktaos_JSS_Final.xlsx")
DATA = os.path.join(BASE, "JSS1 results.xlsx")
OUT = os.path.join(BASE, "report_cards")
CLASS_FOLDER = "JSS 1"

CLASS_NAME = "JSS 1"
NEXT_CLASS = "JSS 2"
N_CLASS = 13           # number of students (filled per data at runtime)

# (data subject prefix, template row, new label or None to keep template label)
SUBJECTS = [
    ("English",            28, None),
    ("Mathematics",        29, None),
    ("Security Education", 30, "Security Education"),
    ("Basic Science",      31, None),
    ("Social Studies",     32, None),
    ("Civic Education",    33, None),
    ("CCA",                34, None),
    ("Agricultural Science", 35, None),
    ("Physical Education", 36, None),
    ("History",            37, "History"),
    ("Home Economics",     38, None),
    ("Basic Technology",   39, None),
    ("Computer Studies",   40, None),
]
UNUSED_ROWS = [41, 42]   # CRS / IRS, French -> cleared
SUBJ_ROWS = [r for _, r, _ in SUBJECTS]
N_SUBJECTS = len(SUBJECTS)

PROT = ["xl/media/image1.jpeg", "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/worksheets/_rels/sheet1.xml.rels"]

GRADE_TABLE = [(75,"A1"),(70,"B2"),(65,"B3"),(60,"C4"),(55,"C5"),
               (50,"C6"),(45,"D7"),(40,"E8"),(0,"F9")]
GPA_MAP = {"A1":4.0,"B2":3.5,"B3":3.0,"C4":2.5,"C5":2.0,
           "C6":1.5,"D7":1.0,"E8":0.5,"F9":0.0}
ORD = ["","1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th",
       "11th","12th","13th","14th","15th","16th","17th","18th"]

THIN = Border(left=Side(style="thin"),right=Side(style="thin"),
              top=Side(style="thin"),bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT = Alignment(horizontal="left",vertical="center",wrap_text=True)
F9 = Font(name="Arial", size=9)
F9B = Font(name="Arial", size=9, bold=True)
F11B = Font(name="Arial", size=11, bold=True)
F12B = Font(name="Arial", size=12, bold=True)
F14G = Font(name="Arial", size=14, bold=True, color="006400")


def grade(t):
    t = max(0, min(100, t))
    for lo, g in GRADE_TABLE:
        if lo <= t: return g
    return "F9"

def ordinal(r):
    return ORD[r] if 1 <= r < len(ORD) else str(r)

def slug(name):
    return re.sub(r"[^\w]+", "_", name).strip("_")


def read_students():
    wb = load_workbook(DATA, data_only=True)
    ws = wb["Scores"]
    hd = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    students = []
    for r in range(2, ws.max_row+1):
        nm = ws.cell(row=r, column=hd["Student Name"]).value
        if not nm or not str(nm).strip():
            continue
        s = {"n": str(nm).strip(), "subj": {}}
        for subj, _, _ in SUBJECTS:
            t = ws.cell(row=r, column=hd[f"{subj}_Test"]).value or 0
            e = ws.cell(row=r, column=hd[f"{subj}_Exam"]).value or 0
            s["subj"][subj] = {"test": int(t), "exam": int(e),
                               "total": int(t) + int(e)}
        s["total"] = sum(v["total"] for v in s["subj"].values())
        students.append(s)
    return students


def rank_map(vals):
    order = sorted(vals.items(), key=lambda kv: -kv[1])
    rk, prev, ri = {}, None, 0
    for i, (n, v) in enumerate(order):
        if prev is not None and v < prev:
            ri = i
        rk[n] = ri + 1
        prev = v
    return rk


def save(path, modify):
    shutil.copy2(TEMPLATE, path)
    wb = load_workbook(path)
    modify(wb)
    tmp = path + ".tmp"
    wb.save(tmp); wb.close()
    with zipfile.ZipFile(tmp) as z:
        o = {n: z.read(n) for n in z.namelist()}
    with zipfile.ZipFile(TEMPLATE) as z:
        for n in PROT:
            if n in z.namelist():
                o[n] = z.read(n)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n in sorted(o):
            z.writestr(n, o[n])
    os.remove(tmp)


def build_card(s, class_rank, subject_ranks, n_class):
    name = s["n"]
    path = os.path.join(OUT, CLASS_FOLDER, "Report_Card_%s.xlsx" % slug(name))
    total_marks = s["total"]
    avg = round(total_marks / N_SUBJECTS, 1)
    pct = round(total_marks / (N_SUBJECTS * 100) * 100, 2)
    overall = grade(round(pct))
    gpa = GPA_MAP.get(overall, 0.0)
    credit = sum(1 for v in s["subj"].values()
                 if grade(v["total"]) in ("A1","B2","B3","C4","C5","C6"))
    pos = ordinal(class_rank)

    def mod(wb):
        ws = wb["Report Sheet"]
        # student info
        ws["D16"].value = name; ws["D16"].font = F12B; ws["D16"].alignment = LEFT
        ws["D20"].value = CLASS_NAME; ws["D20"].font = F12B; ws["D20"].alignment = LEFT
        ws["M19"].value = n_class; ws["M19"].font = F12B; ws["M19"].alignment = CENTER
        ws["M20"].value = pos;     ws["M20"].font = F12B; ws["M20"].alignment = CENTER
        ws["M21"].value = n_class; ws["M21"].font = F12B; ws["M21"].alignment = CENTER

        # relabel reassigned subject rows
        for subj, row, label in SUBJECTS:
            if label:
                c = ws.cell(row=row, column=2)
                c.value = label; c.font = F9; c.alignment = LEFT; c.border = THIN

        # clear unused subject rows (CRS/IRS, French)
        for row in UNUSED_ROWS:
            for col in (1, 2, 7, 8, 10, 12, 14, 16):
                try: ws.cell(row=row, column=col).value = ""
                except Exception: pass

        # subjects
        for subj, row, _ in SUBJECTS:
            v = s["subj"][subj]
            t, e = v["test"], v["exam"]
            g = round(t / 2); h = t - g
            ws.cell(row=row, column=7).value = g
            ws.cell(row=row, column=8).value = h
            ws.cell(row=row, column=10).value = e
            for c in (7, 8, 10):
                cell = ws.cell(row=row, column=c)
                cell.font = F9; cell.alignment = CENTER; cell.border = THIN
            sr = subject_ranks[subj].get(name, 0)
            ws.cell(row=row, column=16).value = ordinal(sr)
            ws.cell(row=row, column=16).font = F9
            ws.cell(row=row, column=16).alignment = CENTER
            ws.cell(row=row, column=16).border = THIN

        # left summary (rows 43-50, col G)
        summary = {43: total_marks, 44: avg, 45: "%s%%" % pct, 46: overall,
                   47: pos, 48: n_class, 49: credit, 50: gpa}
        for r, val in summary.items():
            ws.cell(row=r, column=7).value = val
            ws.cell(row=r, column=7).font = F11B

        # right performance box (X43 Total Number, X44 Credit, X45 GPA)
        for r, val in [(43, n_class), (44, credit), (45, gpa)]:
            ws.cell(row=r, column=24).value = val
            ws.cell(row=r, column=24).font = F11B

        # promotion + next class
        if ws["T60"].value == "PROMOTED":
            ws["T60"].value = "\u2714  PROMOTED"
            ws["T60"].font = F14G; ws["T60"].alignment = CENTER
        ws["V64"].value = NEXT_CLASS
        ws["V64"].font = F12B; ws["V64"].alignment = CENTER

    save(path, mod)
    return path


def main():
    os.makedirs(os.path.join(OUT, CLASS_FOLDER), exist_ok=True)
    students = read_students()
    if not students:
        print("No students found in JSS1 results.xlsx"); return
    n_class = len(students)
    class_rank = rank_map({s["n"]: s["total"] for s in students})
    subject_ranks = {subj: rank_map({s["n"]: s["subj"][subj]["total"]
                                     for s in students})
                     for subj, _, _ in SUBJECTS}
    students.sort(key=lambda s: -s["total"])

    print("\n\U0001F4DA JSS 1  (%d students, %d subjects)" % (n_class, N_SUBJECTS))
    print("%-4s %-6s %-7s %s" % ("Pos", "Total", "Avg", "Student"))
    for s in students:
        print("%-4s %-6d %-7s %s" % (ordinal(class_rank[s["n"]]), s["total"],
              round(s["total"]/N_SUBJECTS,1), s["n"]))
        build_card(s, class_rank[s["n"]], subject_ranks, n_class)
        print("        \u2705 %s" % s["n"])


if __name__ == "__main__":
    main()
