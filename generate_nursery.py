#!/usr/bin/env python3
"""
Hiktaos Nursery 1 & Nursery 2 Report Card Generator.

Data sources : Nursery1.xlsx, Nursery2.xlsx (sheet 'Scores'; columns
               '<Subject>_CA', '<Subject>_Exam', '<Subject>_Total')
Templates    : Hiktaos_Nursery1_Final.xlsx, Hiktaos_Nursery2_Final.xlsx
               (logo at xl/media/image1.jpeg)

Template score columns (subject rows 28-40):
    G = CA 1, H = Assignment, J = Examination,
    L = Total -> formula =SUM(G,H,J), N = Grade (formula), P = Position.

CA is split across the two CA boxes and Exam goes in the exam box:
    G = round(CA / 2), H = CA - G, J = Exam.

Per user instruction, the card total honours the recorded '_Total' value:
where CA + Exam != Total, CA is set to (Total - Exam) so the card shows the
recorded Total. (For all consistent cells Total - Exam already equals CA.)

Both classes have 13 subjects that map cleanly to template rows 28-40 in the
same order, so no relabelling is needed.
"""
import os, re, shutil, zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "report_cards")

PROT = ["xl/media/image1.jpeg", "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/worksheets/_rels/sheet1.xml.rels"]

GRADE_TABLE = [(75,"A1"),(70,"B2"),(65,"B3"),(60,"C4"),(55,"C5"),
               (50,"C6"),(45,"D7"),(40,"E8"),(0,"F9")]
GPA_MAP = {"A1":4.0,"B2":3.5,"B3":3.0,"C4":2.5,"C5":2.0,
           "C6":1.5,"D7":1.0,"E8":0.5,"F9":0.0}
ORD = ["","1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th",
       "11th","12th","13th","14th","15th","16th","17th","18th","19th","20th"]

THIN = Border(left=Side(style="thin"),right=Side(style="thin"),
              top=Side(style="thin"),bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT = Alignment(horizontal="left",vertical="center",wrap_text=True)
F9 = Font(name="Arial", size=9)
F11B = Font(name="Arial", size=11, bold=True)
F12B = Font(name="Arial", size=12, bold=True)
F14G = Font(name="Arial", size=14, bold=True, color="006400")

def grade(t):
    t = max(0, min(100, t))
    for lo, g in GRADE_TABLE:
        if lo <= t: return g
    return "F9"

def ordinal(r):
    return ORD[r] if 1 <= r < len(ORD) else str(r)

def slug(name):
    return re.sub(r"[^\w]+", "_", name).strip("_")

# (data subject prefix, template row)
N1_SUBJECTS = [("Rhymes",28),("Food Nutrition",29),("Creative Art",30),
    ("Handwriting",31),("Health Habit",32),("Quantitative",33),("Verbal",34),
    ("PHE",35),("Civic",36),("Basic Science",37),("Social Habit",38),
    ("Number Works",39),("Letter Works",40)]
N2_SUBJECTS = [("Number Works",28),("Letter Works",29),("Quantitative",30),
    ("Verbal",31),("PHE",32),("Civic Education",33),("Basic Science",34),
    ("Social Habit",35),("Creative Art",36),("Handwriting",37),
    ("Health Habit",38),("Rhymes",39),("Food & Nutrition",40)]

KG_SUBJECTS = [("English",28),("Math",29),("Quantitative",30),("Verbal",31),
    ("Health Habit",32),("Basic Science",33),("CCA",34),("Food & Nut",35),
    ("Handwriting",36),("Civic",37),("Social Habit",38),("PHE",39)]

CLASSES = [
    {"name":"Nursery 1","next":"Nursery 2","data":"Nursery1.xlsx",
     "template":"Hiktaos_Nursery1_Final.xlsx","summary_start":41,"subjects":N1_SUBJECTS},
    {"name":"Nursery 2","next":"KG","data":"Nursery2.xlsx",
     "template":"Hiktaos_Nursery2_Final.xlsx","summary_start":41,"subjects":N2_SUBJECTS},
    {"name":"KG","next":"Primary 1","data":"KG.xlsx",
     "template":"Hiktaos_KG_Final.xlsx","summary_start":40,"subjects":KG_SUBJECTS},
]


def read_students(cfg):
    wb = load_workbook(os.path.join(BASE, cfg["data"]), data_only=True)
    ws = wb["Scores"]
    hd = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    students = []
    for r in range(2, ws.max_row+1):
        nm = ws.cell(row=r, column=hd["Student Name"]).value
        if not nm or not str(nm).strip():
            continue
        s = {"n": str(nm).strip(), "subj": {}}
        for subj, _ in cfg["subjects"]:
            ca = ws.cell(row=r, column=hd[f"{subj}_CA"]).value or 0
            ex = ws.cell(row=r, column=hd[f"{subj}_Exam"]).value or 0
            tot = ws.cell(row=r, column=hd[f"{subj}_Total"]).value or 0
            tot = int(tot)
            ca_eff = int(tot) - int(ex)          # honour recorded Total
            s["subj"][subj] = {"ca": ca_eff, "exam": int(ex), "total": tot}
        s["total"] = sum(v["total"] for v in s["subj"].values())
        students.append(s)
    return students


def rank_map(vals):
    order = sorted(vals.items(), key=lambda kv: -kv[1])
    rk, prev, ri = {}, None, 0
    for i, (n, v) in enumerate(order):
        if prev is not None and v < prev:
            ri = i
        rk[n] = ri + 1
        prev = v
    return rk


def save(path, template, modify):
    shutil.copy2(template, path)
    wb = load_workbook(path)
    modify(wb)
    tmp = path + ".tmp"
    wb.save(tmp); wb.close()
    with zipfile.ZipFile(tmp) as z:
        o = {n: z.read(n) for n in z.namelist()}
    with zipfile.ZipFile(template) as z:
        for n in PROT:
            if n in z.namelist():
                o[n] = z.read(n)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n in sorted(o):
            z.writestr(n, o[n])
    os.remove(tmp)


def build_card(cfg, s, class_rank, subject_ranks, n_class):
    name = s["n"]
    nsubj = len(cfg["subjects"])
    out_dir = os.path.join(OUT, cfg["name"])
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "Report_Card_%s.xlsx" % slug(name))
    total_marks = s["total"]
    avg = round(total_marks / nsubj, 1)
    pct = round(total_marks / (nsubj * 100) * 100, 2)
    overall = grade(round(pct))
    gpa = GPA_MAP.get(overall, 0.0)
    credit = sum(1 for v in s["subj"].values()
                 if grade(v["total"]) in ("A1","B2","B3","C4","C5","C6"))
    pos = ordinal(class_rank)

    def mod(wb):
        ws = wb["Report Sheet"]
        ws["D16"].value = name; ws["D16"].font = F12B; ws["D16"].alignment = LEFT
        ws["D20"].value = cfg["name"]; ws["D20"].font = F12B; ws["D20"].alignment = LEFT
        ws["M19"].value = n_class; ws["M19"].font = F12B; ws["M19"].alignment = CENTER
        ws["M20"].value = pos; ws["M20"].font = F12B; ws["M20"].alignment = CENTER
        ws["M21"].value = n_class; ws["M21"].font = F12B; ws["M21"].alignment = CENTER

        for subj, row in cfg["subjects"]:
            v = s["subj"][subj]
            ca, ex = v["ca"], v["exam"]
            g = round(ca / 2); h = ca - g
            ws.cell(row=row, column=7).value = g
            ws.cell(row=row, column=8).value = h
            ws.cell(row=row, column=10).value = ex
            for c in (7, 8, 10):
                cell = ws.cell(row=row, column=c)
                cell.font = F9; cell.alignment = CENTER; cell.border = THIN
            sr = subject_ranks[subj].get(name, 0)
            ws.cell(row=row, column=16).value = ordinal(sr)
            ws.cell(row=row, column=16).font = F9
            ws.cell(row=row, column=16).alignment = CENTER
            ws.cell(row=row, column=16).border = THIN

        # left summary (col G) starting at class's summary_start row
        ss = cfg["summary_start"]
        summary = {ss: total_marks, ss+1: avg, ss+2: "%s%%" % pct, ss+3: overall,
                   ss+4: pos, ss+5: n_class, ss+6: credit, ss+7: gpa}
        for r, val in summary.items():
            ws.cell(row=r, column=7).value = val
            ws.cell(row=r, column=7).font = F11B

        # right performance box rows 38-45 (col X=24)
        for r, val in [(38,total_marks),(39,avg),(40,"%s%%"%pct),(41,overall),
                       (42,pos),(43,n_class),(44,credit),(45,gpa)]:
            ws.cell(row=r, column=24).value = val
            ws.cell(row=r, column=24).font = F11B

        if ws["T60"].value == "PROMOTED":
            ws["T60"].value = "\u2714  PROMOTED"
            ws["T60"].font = F14G; ws["T60"].alignment = CENTER
        ws["V64"].value = cfg["next"]
        ws["V64"].font = F12B; ws["V64"].alignment = CENTER

    save(path, os.path.join(BASE, cfg["template"]), mod)
    return path


def run_class(cfg):
    students = read_students(cfg)
    if not students:
        print("No students in", cfg["data"]); return
    n_class = len(students)
    nsubj = len(cfg["subjects"])
    class_rank = rank_map({s["n"]: s["total"] for s in students})
    subject_ranks = {subj: rank_map({s["n"]: s["subj"][subj]["total"]
                                     for s in students})
                     for subj, _ in cfg["subjects"]}
    students.sort(key=lambda s: -s["total"])

    print("\n\U0001F4DA %s  (%d students, %d subjects) -> next: %s"
          % (cfg["name"], n_class, nsubj, cfg["next"]))
    print("%-4s %-6s %-7s %s" % ("Pos", "Total", "Avg", "Student"))
    for s in students:
        print("%-4s %-6d %-7s %s" % (ordinal(class_rank[s["n"]]), s["total"],
              round(s["total"]/nsubj,1), s["n"]))
        build_card(cfg, s, class_rank[s["n"]], subject_ranks, n_class)
        print("        \u2705 %s" % s["n"])


def main():
    os.makedirs(OUT, exist_ok=True)
    for cfg in CLASSES:
        run_class(cfg)


if __name__ == "__main__":
    main()
